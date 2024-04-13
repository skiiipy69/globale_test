import openpyxl
import os

def search_and_save_sheets(workbook_path, search_term):
    # Load the workbook
    workbook = openpyxl.load_workbook(workbook_path)

    # Create a new workbook for output
    output_workbook = openpyxl.Workbook()

    # Create a list to store the found sheets
    found_sheets = []

    # Loop through all worksheets in the workbook
    for sheet in workbook.worksheets:
        # Check if the search_term exists in the current worksheet
        if any(search_term.lower() in str(cell.value).lower() for row in sheet.iter_rows() for cell in row):
            found_sheets.append(sheet)

    # Save the found sheets to separate sheets in the output workbook
    if found_sheets:
        for index, sheet in enumerate(found_sheets, start=1):
            new_sheet = output_workbook.create_sheet(title=f"{search_term}_{index}")
            for row in sheet.iter_rows():
                new_sheet.append([cell.value for cell in row])

        save_dir = os.path.dirname(workbook_path)
        save_path = os.path.join(save_dir, f"{search_term}_output.xlsx")
        output_workbook.save(save_path)
        print(f"Found and saved {len(found_sheets)} sheets containing '{search_term}' to {save_path}")
    else:
        print(f"No sheets found containing '{search_term}'.")

# Prompt the user for the search term
search_term = input("Enter the name to search for: ")

# Set the workbook path
workbook_path = "scraped_data_with_sheets.xlsx"

# Call the search_and_save_sheets function
search_and_save_sheets(workbook_path, search_term)