import openpyxl

def nouvelle_operation(ws):
    ws["B7"].value = None
    ws["B8"].value = None
    ws["B10"].value = None
    for row in ws.iter_rows(min_row=13, max_row=99, min_col=2, max_col=11):
        for cell in row:
            cell.value = None

def calculer_fourchette(ws):
    estmation = ws["B6"].value
    categories = ws["B5"].value
    
    if categories == "Travaux":
        seuil_excessif = estmation * 1.2
        seuil_bas = estmation * 0.8
        return
    
    ws["B7"].value = seuil_excessif
    ws["B8"].value = seuil_bas

def offres_ecartes(ws):
    taille = ws.max_row - 12
    
    soumissionnaires = [ws.cell(row=i+13, column=1).value for i in range(taille)]
    offres = [ws.cell(row=i+13, column=2).value for i in range(taille)]
    
    estmation = ws["B6"].value
    seuil_excessif = ws["B7"].value
    seuil_bas = ws["B8"].value
    
    for i in range(taille):
        if offres[i] is not None and offres[i] != 0:
            if offres[i] > seuil_excessif or offres[i] < seuil_bas:
                ws.cell(row=i+13, column=4).value = soumissionnaires[i]
                ws.cell(row=i+13, column=5).value = offres[i]

def prix_de_reference(ws):
    taille = ws.max_row - 12
    
    offres = [ws.cell(row=i+13, column=2).value for i in range(taille)]
    
    estmation = ws["B6"].value
    seuil_excessif = ws["B7"].value
    seuil_bas = ws["B8"].value
    
    somme_offres = sum(offre for offre in offres if seuil_bas <= offre <= seuil_excessif)
    nbr_offres_valide = sum(1 for offre in offres if seuil_bas <= offre <= seuil_excessif)
    
    if nbr_offres_valide == 0:
        print("AOO Infructueux ....!")
        return
    
    px_reference = ((somme_offres / nbr_offres_valide) + estmation) / 2
    ws["B10"].value = px_reference

def offres_proche_par_defaut(ws):
    taille = ws.max_row - 12
    
    soumissionnaires = [ws.cell(row=i+13, column=1).value for i in range(taille)]
    offres = [ws.cell(row=i+13, column=2).value for i in range(taille)]
    
    estmation = ws["B6"].value
    seuil_excessif = ws["B7"].value
    seuil_bas = ws["B8"].value
    px_reference = ws["B10"].value
    
    if px_reference is not None:
        for i in range(taille):
            ecart_final = None
            for p in range(i, taille):
                if offres[p] is not None and offres[p] > px_reference:
                    continue
                ecart_temp = abs(px_reference - offres[p])
                if ecart_final is not None and ecart_temp < ecart_final:
                    ecart_final = ecart_temp
                    concurrent = soumissionnaires[p]
                    offre = offres[p]
                    soumissionnaires[p] = soumissionnaires[i]
                    offres[p] = offres[i]
                    soumissionnaires[i] = concurrent
                    offres[i] = offre
                elif ecart_final is None:
                    ecart_final = ecart_temp
                    concurrent = soumissionnaires[p]
                    offre = offres[p]
                    soumissionnaires[p] = soumissionnaires[i]
                    offres[p] = offres[i]
                    soumissionnaires[i] = concurrent
                    offres[i] = offre
            if offres[i] is not None and seuil_bas <= offres[i] <= px_reference:
                ws.cell(row=i+13, column=7).value = soumissionnaires[i]
                ws.cell(row=i+13, column=8).value = offres[i]

def offres_proche_par_exces(ws):
    taille = ws.max_row - 12
    
    soumissionnaires = [ws.cell(row=i+13, column=1).value for i in range(taille)]
    offres = [ws.cell(row=i+13, column=2).value for i in range(taille)]
    
    estmation = ws["B6"].value
    seuil_excessif = ws["B7"].value
    seuil_bas = ws["B8"].value
    px_reference = ws["B10"].value
    
    if px_reference is not None:
        for i in range(taille):
            ecart_final = None
            for p in range(i, taille):
                if offres[p] is not None and offres[p] <= px_reference:
                    continue
                ecart_temp = abs(px_reference - offres[p])
                if ecart_final is not None and ecart_temp < ecart_final:
                    ecart_final = ecart_temp
                    concurrent = soumissionnaires[p]
                    offre = offres[p]
                    soumissionnaires[p] = soumissionnaires[i]
                    offres[p] = offres[i]
                    soumissionnaires[i] = concurrent
                    offres[i] = offre
                elif ecart_final is None:
                    ecart_final = ecart_temp
                    concurrent = soumissionnaires[p]
                    offre = offres[p]
                    soumissionnaires[p] = soumissionnaires[i]
                    offres[p] = offres[i]
                    soumissionnaires[i] = concurrent
                    offres[i] = offre
            if offres[i] is not None and seuil_excessif >= offres[i] > px_reference:
                ws.cell(row=i+13, column=10).value = soumissionnaires[i]
                ws.cell(row=i+13, column=11).value = offres[i]

# Example of usage:
wb = openpyxl.load_workbook("your_excel_file.xlsx")
ws = wb.active

nouvelle_operation(ws)
calculer_fourchette(ws)
offres_ecartes(ws)
prix_de_reference(ws
